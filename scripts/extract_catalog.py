# -*- coding: utf-8 -*-
"""
Extrae datos del catálogo Isuzu desde archivos Excel de cotización.
Salida: catalog_data.json
"""

import json
import os
import re
from pathlib import Path

import openpyxl


def _normalize_modelo_key(modelo: str) -> str:
    """Normaliza clave para evitar duplicados Forward/forward, ELF/elf."""
    if not modelo:
        return modelo
    s = str(modelo).strip()
    if s.upper().startswith("FORWARD"):
        return "FORWARD " + s[7:].strip()
    if s.upper().startswith("ELF"):
        return "ELF " + s[3:].strip()
    return s


def clean_value(v):
    """Convierte valores a tipo JSON-serializable."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:  # NaN
            return None
        return v
    s = str(v).strip()
    if s == "" or s.upper() in ("#DIV/0!", "#N/A", "#REF!", "#VALUE!"):
        return None
    return s


def find_file(base_dir, pattern_func):
    """Encuentra archivo por patrón (función que recibe nombre y retorna bool)."""
    for f in os.listdir(base_dir):
        if f.endswith(".xlsx") and pattern_func(f):
            return os.path.join(base_dir, f)
    return None


def extract_datos1(wb, lineas):
    """Extrae precios y catálogo de DATOS1."""
    if "DATOS1" not in wb.sheetnames:
        return
    ws = wb["DATOS1"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return
    # Headers: CATALOGO, AÑO MODELO, AÑO MODELO, PRECIO 2025, PRECIO 2026
    for row in rows[1:]:
        modelo = clean_value(row[0])
        if not modelo:
            continue
        modelo = _normalize_modelo_key(modelo)
        precio_2025 = clean_value(row[3])
        precio_2026 = clean_value(row[4])
        lineas[modelo] = lineas.get(modelo) or {}
        lineas[modelo]["precio_2025"] = precio_2025
        lineas[modelo]["precio_2026"] = precio_2026


def extract_datos2(wb, lineas, linea="ELF"):
    """Extrae datos de hoja DATOS 2. Solo procesa modelos que pertenecen a esta línea."""
    if "DATOS 2" not in wb.sheetnames:
        return
    ws = wb["DATOS 2"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return
    prefix = "FORWARD" if linea == "Forward" else "ELF"
    headers = [str(h).strip() if h else "" for h in rows[0]]
    col_map = {h.upper().replace(" ", "_"): i for i, h in enumerate(headers) if h}
    col_map["IB_PIE"] = col_map.get("IB-PIE_@_RPM", col_map.get("IB-PIE", -1))
    for i, h in enumerate(headers):
        if "pie" in h.lower() or "torque" in h.lower():
            col_map["TORQUE_COL"] = i
            break
    if "TORQUE_COL" not in col_map:
        for i, h in enumerate(headers):
            if h and ("pie" in str(h).lower() or "lb" in str(h).lower()):
                col_map["TORQUE_COL"] = i
                break

    def get_col(name, alt_names=None):
        for n in [name] + (alt_names or []):
            for k, idx in col_map.items():
                if n.upper() in k.replace("-", "_") or k == n.upper():
                    return idx
        return -1

    for row in rows[1:]:
        modelo = clean_value(row[0])
        if not modelo:
            continue
        # Solo procesar modelos de esta línea (ELF vs Forward)
        if prefix == "ELF" and "ELF" not in str(modelo).upper():
            continue
        if prefix == "FORWARD" and "FORWARD" not in str(modelo).upper():
            continue
        modelo = _normalize_modelo_key(modelo)
        lineas[modelo] = lineas.get(modelo) or {}
        lineas[modelo]["linea"] = linea
        lineas[modelo]["modelo"] = modelo
        lineas[modelo]["ano_modelo"] = clean_value(row[1]) if len(row) > 1 else None
        lineas[modelo]["precio"] = clean_value(row[2]) if len(row) > 2 else None
        lineas[modelo]["capacidad_carga"] = clean_value(row[3]) if len(row) > 3 else None
        lineas[modelo]["largo_chasis"] = clean_value(row[4]) if len(row) > 4 else None
        lineas[modelo]["alto_camion"] = clean_value(row[5]) if len(row) > 5 else None
        lineas[modelo]["ancho_cabina"] = clean_value(row[6]) if len(row) > 6 else None
        lineas[modelo]["largo_aplicacion"] = clean_value(row[7]) if len(row) > 7 else None
        lineas[modelo]["alto_aplicacion"] = clean_value(row[8]) if len(row) > 8 else None
        lineas[modelo]["ancho_aplicacion"] = clean_value(row[9]) if len(row) > 9 else None
        lineas[modelo]["distancia_entre_ejes"] = clean_value(row[10]) if len(row) > 10 else None
        lineas[modelo]["pvb"] = clean_value(row[11]) if len(row) > 11 else None
        lineas[modelo]["motor"] = clean_value(row[12]) if len(row) > 12 else None
        lineas[modelo]["tecnologia"] = clean_value(row[13]) if len(row) > 13 else None
        lineas[modelo]["garantia"] = clean_value(row[14]) if len(row) > 14 else None
        lineas[modelo]["equipo"] = clean_value(row[15]) if len(row) > 15 else None
        lineas[modelo]["frenos"] = clean_value(row[16]) if len(row) > 16 else None
        lineas[modelo]["hp"] = clean_value(row[17]) if len(row) > 17 else None
        # Torque: col 18 (ib-pie @ rpm)
        lineas[modelo]["torque"] = clean_value(row[18]) if len(row) > 18 else None
        lineas[modelo]["km_litro"] = clean_value(row[19]) if len(row) > 19 else None
        lineas[modelo]["llantas"] = clean_value(row[22]) if len(row) > 22 else None


def extract_cubicaje(wb, lineas):
    """Extrae medidas de Cubicaje y Peso."""
    if "Cubicaje y Peso" not in wb.sheetnames:
        return
    ws = wb["Cubicaje y Peso"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 3: headers, Row 4+: data. Col 1 = Modelo
    # Medidas: Largo/Ancho/Alto (exterior, interior carga, etc.)
    for row in rows[3:]:
        modelo = clean_value(row[1])
        if not modelo:
            continue
        modelo = _normalize_modelo_key(modelo)
        lineas[modelo] = lineas.get(modelo) or {}
        cubicaje = {}
        # Exterior: col 2=Largo, 3=Ancho, 4=Alto
        cubicaje["exterior"] = {
            "largo_cm": clean_value(row[2]),
            "ancho_cm": clean_value(row[3]),
            "alto_cm": clean_value(row[4]),
        }
        # Interior carga 1: col 6,7,8
        cubicaje["interior_carga_1"] = {
            "largo_cm": clean_value(row[6]),
            "ancho_cm": clean_value(row[7]),
            "alto_cm": clean_value(row[8]),
        }
        # Producto y peso: col 10,11,12,13,14,15
        cubicaje["producto_peso"] = {
            "largo_cm": clean_value(row[10]),
            "ancho_cm": clean_value(row[11]),
            "alto_cm": clean_value(row[12]),
            "kg_por_cu": clean_value(row[13]),
            "habilitados": clean_value(row[14]),
            "peso_total": clean_value(row[15]),
        }
        # Exterior carga: col 18,19,20
        cubicaje["exterior_carga"] = {
            "largo_cm": clean_value(row[18]),
            "ancho_cm": clean_value(row[19]),
            "alto_cm": clean_value(row[20]),
        }
        # Interior carga 2: col 22,23,24
        cubicaje["interior_carga_2"] = {
            "largo_cm": clean_value(row[22]),
            "ancho_cm": clean_value(row[23]),
            "alto_cm": clean_value(row[24]),
        }
        lineas[modelo]["cubicaje_peso"] = {k: v for k, v in cubicaje.items() if any(v.values())}


def extract_distancia_consumo(wb, lineas):
    """Extrae consumo y distancia de Distancia y Consumo."""
    if "Distancia y Consumo" not in wb.sheetnames:
        return
    ws = wb["Distancia y Consumo"]
    rows = list(ws.iter_rows(values_only=True))
    # Row 3: headers, Row 4+: data. Col 1 = Modelo
    for row in rows[3:]:
        modelo = clean_value(row[1])
        if not modelo:
            continue
        modelo = _normalize_modelo_key(modelo)
        lineas[modelo] = lineas.get(modelo) or {}
        consumo = {
            "capacidad_tanque_litros": clean_value(row[2]),
            "rendimiento_min_km_litro": clean_value(row[3]),
            "rendimiento_promedio_km_litro": clean_value(row[4]),
            "rendimiento_max_km_litro": clean_value(row[5]),
            "distancia_min_km": clean_value(row[6]),
            "distancia_promedio_km": clean_value(row[7]),
            "distancia_max_km": clean_value(row[8]),
            "hp": clean_value(row[9]),
            "torque_lb_pie": clean_value(row[10]),
            "relacion_1ra_velocidad": clean_value(row[11]),
            "relacion_diferencial": clean_value(row[12]),
            "torque_final_lb_pie": clean_value(row[13]),
            "motor_cilindros_litros": clean_value(row[14]),
            "frenos_tipo_rin": clean_value(row[15]),
        }
        lineas[modelo]["distancia_consumo"] = {k: v for k, v in consumo.items() if v is not None}


def main():
    base = Path(__file__).resolve().parent.parent
    lineas = {}

    # 1. ELF - Cotización (usar el que NO tiene (1) si ambos existen)
    elf_files = [
        f for f in os.listdir(base)
        if f.endswith(".xlsx") and "ELF" in f.upper() and "Chasis" in f and "Cabina" in f
    ]
    elf_file = None
    for f in elf_files:
        if "(1)" not in f:
            elf_file = base / f
            break
    if not elf_file and elf_files:
        elf_file = base / elf_files[0]

    if elf_file and elf_file.exists():
        wb = openpyxl.load_workbook(str(elf_file), read_only=True, data_only=True)
        extract_datos1(wb, lineas)
        extract_datos2(wb, lineas, "ELF")
        wb.close()

    # 2. Forward
    fwd_files = [
        f for f in os.listdir(base)
        if f.endswith(".xlsx") and "Forward" in f and "Chasis" in f and "Cabina" in f
    ]
    if fwd_files:
        fwd_file = base / fwd_files[0]
        if fwd_file.exists():
            wb = openpyxl.load_workbook(str(fwd_file), read_only=True, data_only=True)
            extract_datos2(wb, lineas, "Forward")
            wb.close()

    # 3. Cubicaje
    cubic_files = [
        f for f in os.listdir(base)
        if f.endswith(".xlsx") and "Cubicaje" in f
    ]
    if cubic_files:
        cubic_file = base / cubic_files[0]
        if cubic_file.exists():
            wb = openpyxl.load_workbook(str(cubic_file), read_only=True, data_only=True)
            extract_cubicaje(wb, lineas)
            extract_distancia_consumo(wb, lineas)
            wb.close()

    # Estructura final por modelo
    catalog = {
        "modelos": {},
        "fuentes": {
            "elf": elf_file.name if elf_file and elf_file.exists() else None,
            "forward": fwd_files[0] if fwd_files else None,
            "cubicaje": cubic_files[0] if cubic_files else None,
        },
    }

    for modelo, datos in sorted(lineas.items()):
        catalog["modelos"][modelo] = datos

    out_path = base / "catalog_data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, indent=2)

    print(f"Guardado: {out_path}")
    print(f"Modelos extraídos: {len(catalog['modelos'])}")


if __name__ == "__main__":
    main()
